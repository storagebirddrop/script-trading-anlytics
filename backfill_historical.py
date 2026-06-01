#!/usr/bin/env python3
"""
Historical Data Backfill Script
Fetches and backfills historical data from January 1, 2024 to present for all assets.
"""

import ccxt
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
import warnings

warnings.filterwarnings('ignore')

# Configuration (same as main script)
ASSETS = [
    # Crypto assets (Binance)
    'BTC', 'ETH', 'SOL', 'XLM', 'REZ', 'RSR', 'NEAR', 'RENDER', 'ONDO', 'ACH', 'BNB', 'XRP',
    # NASDAQ stocks (Yahoo Finance)
    'MSTR', 'XXI', 'RIOT', 'MARA', 'IREN', 'BMNR', 'HUT', 'WULF', 'HIVE', 'CLSK', 'SLNH',
    # LSE ETFs (Yahoo Finance with .L suffix)
    'MSTY', 'YMST', 'MARY', 'RIOY', 'IREY', 'BMNY'
]

ASSET_CONFIG = {
    # Crypto assets (Binance)
    'BTC': {'source': 'binance', 'symbol': 'BTC/USDT'},
    'ETH': {'source': 'binance', 'symbol': 'ETH/USDT'},
    'SOL': {'source': 'binance', 'symbol': 'SOL/USDT'},
    'XLM': {'source': 'binance', 'symbol': 'XLM/USDT'},
    'REZ': {'source': 'binance', 'symbol': 'REZ/USDT'},
    'RSR': {'source': 'binance', 'symbol': 'RSR/USDT'},
    'NEAR': {'source': 'binance', 'symbol': 'NEAR/USDT'},
    'RENDER': {'source': 'binance', 'symbol': 'RENDER/USDT'},
    'ONDO': {'source': 'binance', 'symbol': 'ONDO/USDT'},
    'ACH': {'source': 'binance', 'symbol': 'ACH/USDT'},
    'BNB': {'source': 'binance', 'symbol': 'BNB/USDT'},
    'XRP': {'source': 'binance', 'symbol': 'XRP/USDT'},
    # NASDAQ stocks (Yahoo Finance)
    'MSTR': {'source': 'yahoo', 'symbol': 'MSTR'},
    'XXI': {'source': 'yahoo', 'symbol': 'XXI'},
    'RIOT': {'source': 'yahoo', 'symbol': 'RIOT'},
    'MARA': {'source': 'yahoo', 'symbol': 'MARA'},
    'IREN': {'source': 'yahoo', 'symbol': 'IREN'},
    'BMNR': {'source': 'yahoo', 'symbol': 'BMNR'},
    'HUT': {'source': 'yahoo', 'symbol': 'HUT'},
    'WULF': {'source': 'yahoo', 'symbol': 'WULF'},
    'HIVE': {'source': 'yahoo', 'symbol': 'HIVE'},
    'CLSK': {'source': 'yahoo', 'symbol': 'CLSK'},
    'SLNH': {'source': 'yahoo', 'symbol': 'SLNH'},
    # LSE ETFs (Yahoo Finance with .L suffix)
    'MSTY': {'source': 'yahoo', 'symbol': 'MSTY.L'},
    'YMST': {'source': 'yahoo', 'symbol': 'YMST.L'},
    'MARY': {'source': 'yahoo', 'symbol': 'MARY.L'},
    'RIOY': {'source': 'yahoo', 'symbol': 'RIOY.L'},
    'IREY': {'source': 'yahoo', 'symbol': 'IREY.L'},
    'BMNY': {'source': 'yahoo', 'symbol': 'BMNY.L'},
}

# Indicator parameters
EMA_PERIOD = 21
ATR_PERIOD = 14
RSI_PERIOD = 14
Z_SCORE_PERIOD = 20

SPREADSHEET_PATH = 'ATR_Tracker_Dashboard.xlsx'
MASTER_CSV_PATH = 'data/master.csv'
HISTORY_CSV_PATH = 'data/history.csv'

# Backfill date range
START_DATE = datetime(2024, 1, 1)
END_DATE = datetime.now()


def fetch_historical_binance(symbol, start_date, end_date, timeframe='1d'):
    """Fetch historical OHLCV data from Binance."""
    exchange = ccxt.binance({'enableRateLimit': True})
    
    # Convert to milliseconds
    since = int(start_date.timestamp() * 1000)
    
    try:
        ohlcv = exchange.fetch_ohlcv(symbol, timeframe, since=since, limit=1000)
        df = pd.DataFrame(ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        df.set_index('timestamp', inplace=True)
        
        # Filter by date range
        df = df[df.index >= start_date]
        df = df[df.index <= end_date]
        
        return df
    except Exception as e:
        print(f"Error fetching {symbol} from Binance: {e}")
        return None


def fetch_historical_yahoo(symbol, start_date, end_date, timeframe='1d'):
    """Fetch historical OHLCV data from Yahoo Finance."""
    interval_map = {'1d': '1d', '1w': '1wk'}
    interval = interval_map.get(timeframe, '1d')
    
    try:
        data = yf.download(symbol, start=start_date, end=end_date, interval=interval, progress=False)
        if data.empty:
            print(f"No data returned for {symbol}")
            return None
        
        # Rename columns to match our format
        data = data.rename(columns={
            'Open': 'open',
            'High': 'high',
            'Low': 'low',
            'Close': 'close',
            'Volume': 'volume'
        })
        
        # Reset index to make timestamp a column
        data = data.reset_index()
        data = data.rename(columns={'Date': 'timestamp'})
        data.set_index('timestamp', inplace=True)
        
        return data
    except Exception as e:
        print(f"Error fetching {symbol} from Yahoo Finance: {e}")
        return None


def calculate_ema(df, period):
    """Calculate Exponential Moving Average."""
    return df['close'].ewm(span=period, adjust=False).mean()


def calculate_atr(df, period):
    """Calculate Average True Range."""
    high = df['high']
    low = df['low']
    close = df['close']
    
    tr1 = high - low
    tr2 = (high - close.shift()).abs()
    tr3 = (low - close.shift()).abs()
    
    tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    atr = tr.ewm(span=period, adjust=False).mean()
    
    return atr


def calculate_rsi(df, period):
    """Calculate Relative Strength Index using Wilder's smoothing."""
    delta = df['close'].diff()
    gain = (delta.where(delta > 0, 0)).ewm(span=period, adjust=False).mean()
    loss = (-delta.where(delta < 0, 0)).ewm(span=period, adjust=False).mean()
    
    rs = gain / loss
    rsi = 100 - (100 / (1 + rs))
    
    return rsi


def calculate_z_score(series, period):
    """Calculate Z-score of a series."""
    rolling_mean = series.rolling(window=period).mean()
    rolling_std = series.rolling(window=period).std()
    z_score = (series - rolling_mean) / rolling_std
    return z_score


def calculate_indicators(df):
    """Calculate all technical indicators."""
    df = df.copy()
    
    # Calculate indicators
    df['EMA21'] = calculate_ema(df, EMA_PERIOD)
    df['ATR'] = calculate_atr(df, ATR_PERIOD)
    df['RSI'] = calculate_rsi(df, RSI_PERIOD)
    df['RSI_Z_Score'] = calculate_z_score(df['RSI'], Z_SCORE_PERIOD)
    
    # Calculate derived metrics
    close = df['close'].squeeze() if isinstance(df['close'], pd.DataFrame) else df['close']
    atr_series = df['ATR'].squeeze() if isinstance(df['ATR'], pd.DataFrame) else df['ATR']
    ema_series = df['EMA21'].squeeze() if isinstance(df['EMA21'], pd.DataFrame) else df['EMA21']
    
    df['ATR_Distance'] = (close - ema_series) / atr_series
    df['Pct_Above_EMA'] = ((close - ema_series) / ema_series) * 100
    
    return df


def get_historical_data(asset, start_date, end_date, timeframe):
    """Fetch and calculate indicators for historical data."""
    config = ASSET_CONFIG.get(asset)
    if not config:
        print(f"Configuration not found for {asset}")
        return None
    
    source = config['source']
    symbol = config['symbol']
    
    if source == 'binance':
        df = fetch_historical_binance(symbol, start_date, end_date, timeframe)
    elif source == 'yahoo':
        df = fetch_historical_yahoo(symbol, start_date, end_date, timeframe)
    else:
        return None
    
    if df is None or df.empty:
        return None
    
    df = calculate_indicators(df)
    
    # Convert to list of dictionaries
    data = []
    for idx, row in df.iterrows():
        def to_scalar(val):
            if isinstance(val, (pd.Series, pd.DataFrame)):
                return float(val.iloc[0] if isinstance(val, pd.Series) else val.iloc[0, 0])
            if pd.isna(val):
                return None
            return float(val)
        
        data.append({
            'Date': idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else str(idx),
            'Asset': asset,
            'Price': to_scalar(row['close']),
            'EMA21': to_scalar(row['EMA21']),
            'ATR': to_scalar(row['ATR']),
            'RSI': to_scalar(row['RSI']),
            'RSI_Z_Score': to_scalar(row['RSI_Z_Score']),
            'ATR_Distance': to_scalar(row['ATR_Distance']),
            'Pct_Above_EMA': to_scalar(row['Pct_Above_EMA']),
            'Timeframe': timeframe
        })
    
    return data


def write_to_master_csv(data):
    """Write data to master.csv (latest snapshot)."""
    df = pd.DataFrame(data)
    df = df.sort_values(['Asset', 'Timeframe', 'Date'])
    df.to_csv(MASTER_CSV_PATH, index=False)
    print(f"Data written to {MASTER_CSV_PATH}")


def write_to_history_csv(data):
    """Write historical data to history.csv (all historical data)."""
    df = pd.DataFrame(data)
    df = df.sort_values(['Date', 'Asset', 'Timeframe'])
    df.to_csv(HISTORY_CSV_PATH, index=False)
    print(f"Data written to {HISTORY_CSV_PATH} (total rows: {len(df)})")


def write_to_sheet(wb, sheet_name, data):
    """Write data to a specific sheet, creating it if it doesn't exist."""
    headers = ['Date', 'Asset', 'Timeframe', 'Price', 'EMA21', 'ATR', 'RSI', 'RSI_Z_Score', 'ATR_Distance', 'Pct_Above_EMA']
    
    if sheet_name in wb.sheetnames:
        # Sheet exists, append data
        ws = wb[sheet_name]
        start_row = ws.max_row + 1
    else:
        # Sheet doesn't exist, create it
        ws = wb.create_sheet(title=sheet_name)
        # Write headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        start_row = 2
    
    # Write data
    for row_num, row_data in enumerate(data, start_row):
        ws.cell(row=row_num, column=1, value=row_data['Date'])
        ws.cell(row=row_num, column=2, value=row_data['Asset'])
        ws.cell(row=row_num, column=3, value=row_data['Timeframe'])
        ws.cell(row=row_num, column=4, value=row_data['Price'])
        ws.cell(row=row_num, column=5, value=row_data['EMA21'])
        ws.cell(row=row_num, column=6, value=row_data['ATR'])
        ws.cell(row=row_num, column=7, value=row_data['RSI'])
        ws.cell(row=row_num, column=8, value=row_data['RSI_Z_Score'])
        ws.cell(row=row_num, column=9, value=row_data['ATR_Distance'])
        ws.cell(row=row_num, column=10, value=row_data['Pct_Above_EMA'])
    
    print(f"  {sheet_name}: Added {len(data)} records (total rows: {ws.max_row})")


def main():
    """Main function to backfill historical data."""
    print("Starting Historical Data Backfill...")
    print(f"Date range: {START_DATE.strftime('%Y-%m-%d')} to {END_DATE.strftime('%Y-%m-%d')}")
    print(f"Total Assets: {len(ASSETS)}")
    print(f"Timeframes: 1d, 1w")
    print()
    
    all_daily_data = []
    all_weekly_data = []
    
    for asset in ASSETS:
        print(f"Processing {asset}...")
        
        # Fetch daily data
        daily_data = get_historical_data(asset, START_DATE, END_DATE, '1d')
        if daily_data:
            all_daily_data.extend(daily_data)
            print(f"  Daily: {len(daily_data)} records")
        else:
            print(f"  Daily: Failed to fetch data")
        
        # Fetch weekly data
        weekly_data = get_historical_data(asset, START_DATE, END_DATE, '1w')
        if weekly_data:
            all_weekly_data.extend(weekly_data)
            print(f"  Weekly: {len(weekly_data)} records")
        else:
            print(f"  Weekly: Failed to fetch data")
    
    print(f"\nTotal records: {len(all_daily_data)} daily, {len(all_weekly_data)} weekly")
    
    # Combine all data with timeframe
    all_data = all_daily_data + all_weekly_data
    
    # Sort data by date before writing
    all_data.sort(key=lambda x: x['Date'])
    
    # Write to CSV files (primary data source)
    write_to_master_csv(all_data)
    write_to_history_csv(all_data)
    
    # Write to Excel spreadsheet (secondary, for GitHub Pages)
    try:
        wb = load_workbook(SPREADSHEET_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    
    # Write all data to single Data sheet
    if all_data:
        write_to_sheet(wb, 'Data', all_data)
    
    # Save workbook
    wb.save(SPREADSHEET_PATH)
    print(f"Data written to Excel spreadsheet (secondary)")
    print(f"\nBackfill complete!")


if __name__ == '__main__':
    main()
