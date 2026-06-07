#!/usr/bin/env python3
"""
Multi-Asset ATR Tracker Script
Fetches current OHLCV data from multiple sources (Binance for crypto,
Yahoo Finance for stocks/ETFs), calculates technical indicators, and
writes to both master.csv and ATR_Tracker_Dashboard.xlsx.
"""

import json
import sys
import warnings
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings('ignore', category=FutureWarning, module='yfinance')
warnings.filterwarnings('ignore', category=FutureWarning, module='pandas')

from trading_utils import (
    ASSETS,
    ASSET_CONFIG,
    MANUAL_DATA,
    TIMEFRAMES,
    SPREADSHEET_PATH,
    MASTER_CSV_PATH,
    MARKET_CAPS_JSON_PATH,
    calculate_indicators,
    fetch_ohlcv_binance,
    fetch_ohlcv_yahoo,
    fetch_ohlcv_ccxt,
    fetch_ohlcv_geckoterminal,
    get_manual_data,
    fetch_market_caps,
)

_PROJECT_ROOT = Path(__file__).resolve().parent
_MASTER_CSV = Path(MASTER_CSV_PATH)
_SPREADSHEET = Path(SPREADSHEET_PATH)

_EXCEL_HEADERS = [
    'Date', 'Asset', 'Timeframe', 'Price', 'EMA21', 'ATR',
    'RSI', 'RSI_Z_Score', 'ATR_Distance', 'Pct_Above_EMA',
    'High', 'Low', 'Volume',
]

_MAX_FAILED_ASSETS = 40  # allow up to 20 assets × 2 timeframes missing (newer tokens + macro may be unavailable)


def get_data(asset, timeframe):
    """Fetch and calculate indicators for the latest bar of an asset."""
    config = ASSET_CONFIG.get(asset)
    if not config:
        print(f"Configuration not found for {asset}")
        return None

    source = config['source']

    if source == 'manual':
        return get_manual_data(asset, timeframe)

    if source == 'binance':
        df = fetch_ohlcv_binance(config['symbol'], timeframe)
    elif source == 'yahoo':
        df = fetch_ohlcv_yahoo(config['symbol'], timeframe)
    elif source == 'ccxt':
        df = fetch_ohlcv_ccxt(config['exchange'], config['symbol'], timeframe)
    elif source == 'geckoterminal':
        df = fetch_ohlcv_geckoterminal(config['network'], config['pool'], timeframe)
    else:
        print(f"Unknown source for {asset}: {source}")
        return None

    if df is None or df.empty:
        return None

    df = calculate_indicators(df)
    latest = df.iloc[-1]

    return {
        'Date': df.index[-1].strftime('%Y-%m-%d') if hasattr(df.index[-1], 'strftime') else str(df.index[-1]),
        'Asset': asset,
        'Price': float(latest['close']),
        'EMA21': float(latest['EMA21']),
        'ATR': float(latest['ATR']),
        'RSI': float(latest['RSI']),
        'RSI_Z_Score': float(latest['RSI_Z_Score']),
        'ATR_Distance': float(latest['ATR_Distance']) if pd.notna(latest['ATR_Distance']) else None,
        'Pct_Above_EMA': float(latest['Pct_Above_EMA']),
        'Timeframe': timeframe,
        'High':   float(df['high'].iloc[-1])   if 'high'   in df.columns and pd.notna(df['high'].iloc[-1])   else None,
        'Low':    float(df['low'].iloc[-1])    if 'low'    in df.columns and pd.notna(df['low'].iloc[-1])    else None,
        'Volume': float(df['volume'].iloc[-1]) if 'volume' in df.columns and pd.notna(df['volume'].iloc[-1]) else None,
    }


def write_to_excel(records):
    """Append new records to ATR_Tracker_Dashboard.xlsx, skipping existing Date+Asset+Timeframe keys."""
    try:
        wb = load_workbook(_SPREADSHEET)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    sheet_name = 'Data'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Build set of existing composite keys
        existing_keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, asset, timeframe = str(row[0]), str(row[1]), str(row[2])
            existing_keys.add(f"{date}|{asset}|{timeframe}")
        start_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(title=sheet_name)
        for col, header in enumerate(_EXCEL_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        existing_keys = set()
        start_row = 2

    new_count = 0
    for row_data in records:
        key = f"{row_data['Date']}|{row_data['Asset']}|{row_data['Timeframe']}"
        if key in existing_keys:
            continue
        ws.cell(row=start_row, column=1, value=row_data['Date'])
        ws.cell(row=start_row, column=2, value=row_data['Asset'])
        ws.cell(row=start_row, column=3, value=row_data['Timeframe'])
        ws.cell(row=start_row, column=4, value=row_data['Price'])
        ws.cell(row=start_row, column=5, value=row_data['EMA21'])
        ws.cell(row=start_row, column=6, value=row_data['ATR'])
        ws.cell(row=start_row, column=7, value=row_data['RSI'])
        ws.cell(row=start_row, column=8, value=row_data['RSI_Z_Score'])
        ws.cell(row=start_row, column=9,  value=row_data['ATR_Distance'])
        ws.cell(row=start_row, column=10, value=row_data['Pct_Above_EMA'])
        ws.cell(row=start_row, column=11, value=row_data.get('High'))
        ws.cell(row=start_row, column=12, value=row_data.get('Low'))
        ws.cell(row=start_row, column=13, value=row_data.get('Volume'))
        existing_keys.add(key)
        start_row += 1
        new_count += 1

    wb.save(_SPREADSHEET)
    print(f"Excel: wrote {new_count} new records to {_SPREADSHEET}")


def main():
    """Fetch current data for all assets and write to CSV + Excel."""
    all_data = []
    failed = 0

    print(f"Fetching data for {len(ASSETS)} assets across {len(TIMEFRAMES)} timeframes...")

    for asset in ASSETS:
        for timeframe in TIMEFRAMES:
            print(f"Processing {asset} ({timeframe})...")
            data = get_data(asset, timeframe)
            if data:
                all_data.append(data)
            else:
                failed += 1
                print(f"  WARNING: failed to fetch {asset} ({timeframe})")

    if all_data:
        _MASTER_CSV.parent.mkdir(parents=True, exist_ok=True)
        df = pd.DataFrame(all_data)
        df.to_csv(_MASTER_CSV, index=False)
        print(f"master.csv: {len(df)} records written to {_MASTER_CSV}")

        write_to_excel(all_data)

    # Fetch and save market cap data from CoinGecko
    market_caps = fetch_market_caps()
    mcap_path = Path(MARKET_CAPS_JSON_PATH)
    mcap_path.parent.mkdir(parents=True, exist_ok=True)
    with open(mcap_path, 'w') as f:
        json.dump({
            'fetched_at': datetime.now(timezone.utc).isoformat(),
            'data': market_caps,
        }, f)

    if failed > _MAX_FAILED_ASSETS:
        print(f"ERROR: {failed} assets failed (threshold {_MAX_FAILED_ASSETS}). Exiting non-zero.")
        sys.exit(1)

    return all_data


if __name__ == "__main__":
    main()
